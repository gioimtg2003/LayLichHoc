# Tool: Lấy lịch học phần 
# Author: Nguyễn Công Giới
# Version: 1.0
# Date: 31/3/2023
import os, sys
import time, html
import tkinter as tk
from tkinter import messagebox
import requests
import openpyxl
# khởi tạo tệp excel
workbook = openpyxl.Workbook()
# Lấy sheet đầu tiên trong tệp
sheet = workbook.active
# # Thiết lập tiêu đề cột
sheet['A1'] = 'Tên môn học'
sheet['B1'] = 'Mã môn học-Tên lớp'
sheet['C1'] = 'Giảng viên'
sheet['D1'] = 'Lịch học'
sheet['E1'] = 'Địa điểm học' 
# Thiết lập độ rộng cột
sheet.column_dimensions['A'].width = 30
sheet.column_dimensions['B'].width = 35
sheet.column_dimensions['C'].width = 24
sheet.column_dimensions['D'].width = 25
sheet.column_dimensions['D'].height = 50
sheet.column_dimensions['E'].width = 15
# đọc file để lấy cookie
with open('account_sv.txt' , 'r',encoding='utf-8') as f:
    cookie = f.read()
    
# Thiết lập header cho request
headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Language': 'en-US,en;q=0.9,vi;q=0.8',
    'Cache-Control': 'max-age=0',
    'Connection': 'keep-alive',
    'Cookie': str(cookie),
    'Referer': 'https://sv.vaa.edu.vn/lich-theo-tuan.html',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'same-origin',
    'Sec-Fetch-User': '?1',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36',
    'sec-ch-ua': '"Google Chrome";v="111", "Not(A:Brand";v="8", "Chromium";v="111"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
}
# Tạo giao diện cửa sổ chính
root = tk.Tk()
root.title("Lấy lịch học phần(vaa)")
root.geometry("500x400")

# check phiên bản
def status():
    response = requests.get('https://nguyenconggioi.000webhostapp.com/state.json').json()
    if(response['code'] == 1):
        pass
    else:
        messagebox.showinfo("Thông báo", "Phiên bản đã hết hạn, vui lòng liên hệ Fb.com/bumbum26.4 để được cập nhật phiên bản mới nhất!")
        sys.exit()

idDotDangKy_ = requests.get('https://sv.vaa.edu.vn/dang-ky-hoc-phan.html', headers=headers).text

# Lấy tên sinh viên
listbox = tk.Listbox(root, width=50, height=16, font=("Times New Roman", 12))
listbox.place(x=20, y=60)
listbox.delete(0, tk.END)  # xóa nội dung trong listbox từ row 0 -> row cuối cùng
seconds = 5
def countdown(index):
        if index > 0:
            msg = f'Bắt đầu lấy dữ liệu sau {index} giây'
            listbox.insert(tk.END, msg)
            root.after(1000, countdown, index-1)
        else:
            listbox.delete(0, tk.END)
            listbox.insert(tk.END, 'Bắt đầu lấy dữ liệu')
            # sau 5 giây thực hiện hàm start
            root.after(ms=5000, func=start_function)
def start():
    status()
    try:
        global TênSV
        TênSV = html.unescape(idDotDangKy_.split('class="user-account-name"')[1].split('title="">')[1].split('</a>')[0])
        lb = tk.Label(root, text="Xin chào: " + TênSV, font=("Times New Roman", 12))
        lb.pack()
        countdown(seconds)
    except:
        messagebox.showinfo("Thông báo", "Cookie Không hợp lệ")
        sys.exit()
    # lấy id đợt đăng ký Kì học mới nhất
def start_function():
    listbox.delete(0, tk.END)
    idDangKy = idDotDangKy_.split('placeholder="Chọn đợt đăng ký"')[1].split('<option value="')[2].split('"')[0]
    MonHocPhanChoDangKy = requests.post('https://sv.vaa.edu.vn/SinhVienDangKy/MonHocPhanChoDangKy', params={'IDDotDangKy':idDangKy, 'IDLoaiDangKy':'1'}, headers=headers).text
    # chỉ số hàng trong sheet đầu tiên của tệp Excel là 2
    row_index = 2
    # Duyệt qua các môn học trong danh sách môn học
    for i in MonHocPhanChoDangKy.split('SelectMonHocChoDangKy'):
        # Nếu là một môn học thì lấy tên môn học và các dư liệu cần thiết để lấy danh sách lớp học phần
        if ('<td class="text-center">' in i):
            tênMônHọc = html.unescape(i.split('text-left">')[1].split('</td>')[0])
            data_mamh = i.split('data-mamh="')[1].split('"')[0]
            data_mahpduohoc = i.split('data-mahpduochoc="')[1].split('"')[0]
            data = {
                'param[IDDotDangKy]':idDangKy,
                'param[MaMonHoc]': data_mamh,
                'param[DSHocPhanDuocHoc]': data_mahpduohoc,
                'param[IsLHPKhongTrungLich]': 'false',
                'param[LoaiDKHP]': '1',
            }
            listbox.insert(tk.END, tênMônHọc)  # In ra tên môn học
            # Lấy danh sách lớp học phần của môn học đó
            lopHocPhanChoDangKy = requests.post('https://sv.vaa.edu.vn/SinhVienDangKy/LopHocPhanChoDangKy', headers=headers, data=data).text
            # Duyệt qua các lớp học phần của môn học đó
            
            for lopHocPhan in lopHocPhanChoDangKy.split('SelectLopHocPhanChoDangKy'):
                # Nếu là một lớp học phần thì lấy mã lớp học phần và các dữ liệu cần thiết để lấy chi tiết lớp học phần
                if ('lớp  học phần</span>: ' in lopHocPhan):
                    # thêm tên môn học vào cột A
                    sheet.cell(row=row_index, column=1).value = tênMônHọc
                    mãHọcPhần = lopHocPhan.split('lớp  học phần</span>:')[1].split('</div')[0]
                    # thêm mã lớp học phần vào cột B
                    sheet.cell(row=row_index, column=2).value = mãHọcPhần
                    #print(lopHocPhan.split('lớp  học phần</span>:')[1].split('</div')[0])
                    listbox.insert(tk.END, mãHọcPhần)  # In ra mã lớp học phần
                    data_guidlhp = lopHocPhan.split('data-guidlhp="')[1].split('"')[0]   
                    data1 = {
                        'param[GuidIDLopHocPhan]' : data_guidlhp,
                    }
                    chiTietLopHocPhanChoDangKy = requests.post('https://sv.vaa.edu.vn/SinhVienDangKy/ChiTietLopHocPhanChoDangKy', headers=headers, data=data1).text
                    
                    # khi không có lịch học sẽ thêm vào cột "note" là "Chưa có lịch học!"
                    if('dkhp-chuacolichhoc' in chiTietLopHocPhanChoDangKy):
                        sheet.cell(row=row_index, column=3).value = "Chưa có lịch học!"
                        sheet.cell(row=row_index, column=4).value = "Chưa có lịch học!"
                        sheet.cell(row=row_index, column=5).value = "Chưa có lịch học!"
                    else:
                        # Duyệt qua các chi tiết lớp học phần của lớp học phần đó
                        for chiTiet in chiTietLopHocPhanChoDangKy.split('SelectChiTietLopHocPhan(this)'):
                            if('data-guididdk' in chiTiet):
                                try:
                                    phòngHọc = html.unescape(chiTiet.split('lang="dkhp-phong">Phòng</span>: <b>')[1].split('</b>')[0].strip())
                                except:
                                    phòngHọc = "Chưa có phòng học!"
                                try:
                                    lịchHọc = html.unescape(chiTiet.split('"dkhp-lichhoc">Lịch học</span>: <b>')[1].split('</b>')[0])
                                except:
                                    lịchHọc = "Chưa có lịch học!"
                                try:
                                    if ('lang="dkhp-gv"' in chiTiet): 
                                        tênGiảngViên = html.unescape(chiTiet.split('lang="dkhp-gv">GV</span>:')[1].split('</div>')[0].strip())
                                    else:
                                        tênGiảngViên = "Chưa có giảng viên!"
                                except:
                                    tênGiảngViên = "Chưa có giảng viên!"
                                # kiểm tra cột lịch học có giá trị hay không, nếu không có thì thêm giá trị vào, nếu có thì thêm xuống dòng
                                if (sheet.cell(row=row_index, column=4).value == None):
                                    sheet.cell(row=row_index, column=4).value = lịchHọc
                                    sheet.row_dimensions[row_index].height = 18
                                else:
                                    sheet.cell(row=row_index, column=4).value += '\n ' + lịchHọc
                                    #set chiều cao cho row đó khi thêm lịch học
                                    sheet.row_dimensions[row_index].height += 18
                                    
                                # kiểm tra cột phòng học có giá trị hay không, nếu không có thì thêm giá trị vào, nếu có thì thêm xuống dòng
                                if(sheet.cell(row=row_index, column=5).value == None):
                                    sheet.cell(row=row_index, column=5).value = phòngHọc
                                else:
                                    sheet.cell(row=row_index, column=5).value += '\n ' + phòngHọc
                                # print(lịchHọc + ' Phòng: ' + phòngHọc + ' Giảng viên: ' + tênGiảngViên)
                                listbox.insert(tk.END, lịchHọc + ' Phòng: ' + phòngHọc + ' Giảng viên: ' + tênGiảngViên)  # In ra lịch học, phòng học, giảng viên
                                root.update()
                        # thêm tên giảng viên vào cột C
                        sheet.cell(row=row_index, column=3).value = tênGiảngViên
                    # Tăng dòng trong excel lên 1
                    row_index += 1
    # Lưu file excel
    workbook.save(f'{TênSV}.xlsx')
    lb_status = tk.Label(root, text=f"Đã lưu file: {TênSV}.xlsx", font=("Times New Roman", 12))
    lb_status.pack()
start_button = tk.Button(root, text="Start", command=start, width=8 ,font=("Times New Roman", 12))
start_button.place(x=20, y=20)
root.mainloop()
